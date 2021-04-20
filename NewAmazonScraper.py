from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
from tkinter import *
import time
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import date
import mmap
from random import randint
import csv
import pandas as pd
import tkinter as tk
from tkinter import filedialog
import tkinter.font as font
import os
import ntpath
import sys


class ScraperUI():
    def __init__(self):
        self.root = tk.Tk()
        self.root.geometry("600x300")
        self.root.title("Select Your Folder To Check Whether The Titles are Listed in Amazon")
        self.folderConfirmation = tk.Label(self.root)
        self.folderConfirmation.pack()
        self.uiFont = font.Font(size=20)
        self.folderDirectory = None
        self.uploadFileButton = None
        self.pincode = None
        self.pincode_box = None
        self.browser = None
        self.GetUploadButton()
        self.SetBrowserPreference()
        self.SetPincodeBox()
        self.SetSavePincodeButton()
        self.root.after(0, self.checkValues)
        self.root.mainloop()

    def GetUploadButton(self):
        self.uploadFileButton = tk.Button(self.root, bg='yellow', text='Select Folder', command=self.UploadAction,
                                          font=self.uiFont)
        self.uploadFileButton.configure(width=600, height=2)
        self.uploadFileButton.pack()

    def SetPincode(self):
        self.pincode = int(self.pincode_box.get("1.0", "end-1c"))
        self.pincode_success = tk.Label(self.root)

        self.pincode_success.config(text="Pincode Saved Successfully:" + str(self.pincode))
        self.pincode_success.pack()

    def SetPincodeBox(self):
        self.pincode_label = tk.Label(self.root)
        self.pincode_label.config(text="Enter Pincode in the text box below and then press Save Pincode Button")
        self.pincode_label.pack()
        self.pincode_box = tk.Text(self.root, height=2,
                                   width=10,
                                   bg="light yellow")
        self.pincode_box.pack()

    def SetSavePincodeButton(self):
        self.savePincodeButton = tk.Button(self.root, bg='yellow', text='Save PinCode', command=self.SetPincode)
        self.savePincodeButton.configure(width=600, height=2)
        self.savePincodeButton.pack()

    def SetBrowserPreference(self):
        self.BrowserVal = StringVar()
        self.firefoxRadioButton = tk.Radiobutton(self.root, text="Firefox", variable=self.BrowserVal, value='Firefox',
                                                 command=self.SetBrowser)
        self.firefoxRadioButton.pack()
        self.chromeRadioButton = tk.Radiobutton(self.root, text="Chrome", variable=self.BrowserVal, value='Chrome',
                                                command=self.SetBrowser)
        self.chromeRadioButton.pack()

    def SetBrowser(self):
        self.browser = str(self.BrowserVal.get())

    def UploadAction(self):
        self.folderDirectory = filedialog.askdirectory()

    def checkValues(self, debug=False):
        if debug:
            print(self.folderDirectory, self.pincode, self.browser)

        if self.folderDirectory is not None:
            self.folderConfirmation.config(
                text="Current Selected Folder Path : " + str(self.folderDirectory))
        else:
            self.folderConfirmation.config(text="PLEASE SELECT A FOLDER")
        if self.folderDirectory and self.pincode and self.browser:
            time.sleep(1)
            self.root.destroy()
        self.root.after(100, self.checkValues)

    def GetPincode(self):
        return self.pincode

    def GetFolderDirectory(self):
        return self.folderDirectory

    def GetBrowserPreference(self):
        return self.browser


class AmazonScraper():
    def __init__(self):
        self.scraperUI = ScraperUI()
        self.pincode = self.scraperUI.GetPincode()
        self.folderDirectory = self.scraperUI.GetFolderDirectory()
        self.browserPreference = self.scraperUI.GetBrowserPreference()
        if self.folderDirectory:
            self.CreateOutputDirectory()
        if self.browserPreference == 'Firefox':
            self.SetupFirefoxDriver()
        else:
            self.SetupChromeDriver()
        if self.pincode and self.folderDirectory and self.browserPreference:
            self.ReadFileNames()

    def SetFileValues(self):
        self.date = date.today().strftime("%d/%m/%Y")
        self.coverType = None
        self.prime = None
        self.ranking = None
        self.buybox = None
        self.buyboxSeller = None
        self.buyboxPrice = 0
        self.shippingPrice = 0
        self.totalISBN = 0
        self.currISBNRecord = 0
        self.reproListing = None
        self.isISBN = True
        self.currentISBN = None
        self.amazonSearchLink = None
        self.firstLink = True
        self.outputFile = None
        self.bookCode = None
        self.pages = 'NA'

    def resetISBNValues(self):
        self.coverType = None
        self.otherSeller = None
        self.prime = 'Not Prime'
        self.ranking = None
        self.buybox = None
        self.buyboxSeller = None
        self.buyboxPrice = 0
        self.shippingPrice = 0
        self.reproListing = None
        self.isISBN = True
        self.amazonSearchLink = None
        self.pages = 'NA'


    def CheckISBN(self):
        for i in self.currentISBN:
            if not i.isdigit() and i != '.':
                self.isISBN = False
                break

    def CreateOutputDirectory(self):

        if self.folderDirectory:
            self.outputDirectory = os.path.join(self.folderDirectory, 'Output')
            try:
                os.mkdir(self.outputDirectory)
            except FileExistsError:
                pass

    def CreateOutputFileName(self, filename):
        self.outputFile = os.path.join(self.outputDirectory, ntpath.basename(filename[:-5] + '_amazon_output.xlsx'))

    def CreateOutputFile(self, filename):
        self.CreateOutputFileName(filename)
        if not os.path.exists(self.outputFile):
            headers_row = ["Date", 'Book Code',"ISBN13", 'Cover Type', "Prime", 'Ranking', "Buybox", 'Buybox Seller',
                           'Buybox Price', "Repro Listing",'Pages']
            output_workbook = Workbook()
            output_worksheet = output_workbook.active
            output_worksheet.append(headers_row)
            output_workbook.save(self.outputFile)

    def CheckPrimeStatus(self):
        try:
            book_box = self.driver.find_element_by_css_selector(
                "div[class='sg-col-4-of-12 sg-col-8-of-16 sg-col-12-of-20 sg-col']")
            if book_box.find_element_by_css_selector(
                    "span[class='aok-relative s-icon-text-medium s-prime']"):
                self.prime = "Prime"
        except:
            pass

    def CheckReproBuyboxSeller(self):
        if self.buybox is None:
            if 'repro' in self.buyboxSeller.lower():
                self.buybox = 'Yes'
                self.reproListing = True
            else:
                self.buybox = 'No'

    def GetPriceFromText(self, text):
        price = ''
        for i in str(text):
            if i.isdigit() or i == '.':
                price += i
        if price:
            return float(price)
        else:
            return 0

    def GetShippingPrice(self):
        if self.prime != 'Prime':
            try:
                shipping = str(
                    self.driver.find_element_by_css_selector('span[class="a-color-base buyboxShippingLabel"]').text)
                self.shippingPrice = self.GetPriceFromText(shipping)
            except NoSuchElementException:
                try:
                    delivery_charge_id = self.driver.find_element_by_id('ddmDeliveryMessage')
                    delivery_charge_text = str(delivery_charge_id.find_element_by_tag_name('a').text)
                    self.shippingPrice = self.GetPriceFromText(delivery_charge_text)
                except NoSuchElementException:
                    pass
                if not self.shippingPrice:
                    self.shippingPrice = 0

    def GetPaperbackHardCoverLinks(self, classname):
        search_results = self.driver.find_elements_by_css_selector(classname)
        for search_result in search_results:
            link = search_result.get_attribute("href")
            if 'Paperback' in search_result.text:
                if [link, 'Paperback'] not in self.allISBNLinks:
                    self.allISBNLinks.append([link, 'Paperback'])
            if 'Hardcover' in search_result.text:
                if [link, 'Hardcover'] not in self.allISBNLinks:
                    self.allISBNLinks.append([link, 'Hardcover'])

    def GetAllValidLinks(self):
        self.allISBNLinks = []
        self.GetPaperbackHardCoverLinks("a[class='a-size-base a-link-normal']")
        self.GetPaperbackHardCoverLinks("a[class='a-size-base a-link-normal a-text-bold']")

    def GetCurrentLink(self):
        for currLink in self.allISBNLinks:
            try:
                self.resetISBNValues()
                self.driver.get(currLink[0])
                self.coverType = currLink[1]
                time.sleep(2)
                if self.firstLink:
                    self.SetPincode()
                    self.firstLink = False
                if self.prime == 'Not Prime':
                    self.GetShippingPrice()
                self.GetISBNRanking()
                self.GetISBNNumber()
                self.GetPages()
                self.GetBuyboxSeller()
                self.CheckReproBuyboxSeller()
                if self.buybox != 'Yes':
                    self.GetOtherBuyingOptionElement()
                    self.CheckReproInMoreBuyingOptions()
                self.WriteRecord()
                self.GetISBNValues()
            except:
                pass


    def GetOtherBuyingOptionElement(self):
        if self.buybox == 'No':
            try:
                self.otherSeller = self.driver.find_element_by_id('mbc-olp-link')
            except NoSuchElementException:
                pass
        elif self.buybox == 'No Buybox':
            try:
                self.otherSeller = self.driver.find_element_by_id('buybox-see-all-buying-choices')
            except NoSuchElementException:
                pass

    def CheckReproInMoreBuyingOptions(self):
        if self.otherSeller:
            # time.sleep(2)
            loadSidePanel = self.otherSeller.find_element_by_tag_name('a')
            loadSidePanel.click()
            time.sleep(3)
            allSellers = self.driver.find_elements_by_css_selector("a[class='a-size-small a-link-normal']")
            for currSeller in allSellers:
                if 'repro' in str(currSeller.text).lower():
                    self.reproListing = True
                    break

    def GetBuyboxSeller(self):
        try:
            sellerIDElement = self.driver.find_element_by_id('sellerProfileTriggerId')
            self.buyboxSeller = str(sellerIDElement.text)
            self.buyboxPrice = str(self.driver.find_element_by_id('soldByThirdParty').text).replace('₹', '').replace(
                ',', '').strip()
            self.buyboxPrice = float(self.buyboxPrice) + float(self.shippingPrice)
        except NoSuchElementException:
            self.buybox = 'No Buybox'

    def GetISBNNumber(self):
        the_details_id = self.driver.find_element_by_id('detailBullets_feature_div')
        span_elements = the_details_id.find_elements_by_class_name('a-list-item')
        isbn = ''
        for span_element in span_elements:
            if 'ISBN-13' in span_element.text:
                temp_isbn = str(span_element.text).split('\n')[0].replace('ISBN-13', '')
                for i in temp_isbn:
                    if i.isdigit():
                        isbn += i
                break
        if isbn and isbn != self.currentISBN:
            self.currentISBN = isbn

    def GetISBNRanking(self):
        the_details_id = self.driver.find_element_by_id('detailBullets_feature_div')
        span_elements = the_details_id.find_elements_by_class_name('a-list-item')
        rank = ""
        for span_element in span_elements:
            if 'Sellers Rank' in span_element.text:
                temp_rank = str(span_element.text).split('\n')[0].replace('See Top 100', '')
                for i in temp_rank:
                    if i.isdigit():
                        rank += i
                break
        if rank:
            self.ranking = rank
        else:
            self.ranking = "NA"

    def GetPages(self):
        the_details_id = self.driver.find_element_by_id('detailBullets_feature_div')
        span_elements = the_details_id.find_elements_by_class_name('a-list-item')
        pages = ''
        for span_element in span_elements:
            if 'pages' in span_element.text:
                for i in span_element.text:
                    if i.isdigit():
                        pages += i
                break
        if pages:
            self.pages = pages
        else:
            self.pages = "NA"

    def SetPincode(self):
        pincodeLabel = self.driver.find_element_by_id('contextualIngressPtLabel')
        pincodeLabel.click()
        time.sleep(2)
        pinApplyID = self.driver.find_element_by_id('GLUXZipInputSection')
        pincodeInputBox = self.driver.find_element_by_css_selector("input[class='GLUX_Full_Width a-declarative']")
        pincodeInputBox.send_keys(self.pincode)
        pincodeSubmitButton = pinApplyID.find_element_by_class_name('a-button-input')
        pincodeSubmitButton.click()
        time.sleep(2)

    def GetISBNValues(self):
        print(str(self.currISBNRecord) + '/' + str(self.totalISBN) + ', Date : ' + str(
            self.date) + ', ISBN13 : ' + str(self.currentISBN) + ', Book Code : ' + str(
            self.bookCode) + ', Cover Type : ' + self.coverType + ', Prime : ' + str(self.prime) + ', Ranking : ' + str(
            self.ranking) + ', Buybox : ' + str(self.buybox) + ', Buybox Seller : ' +
              str(self.buyboxSeller) + ', Buybox Price : ' + str(self.buyboxPrice) + ', Repro Listing : ' + str(
            self.reproListing)  + ', Pages : ' + self.pages)

    def OpenAmazonPage(self):
        self.driver.get(self.amazonSearchLink)

    def ReadCurrentFile(self, filename):
        self.currentExcelFile = pd.read_excel(filename)

    def ReadISBN(self, filename):
        try:
            self.isbnList = self.currentExcelFile['ISBN13']
            self.totalISBN = len(self.isbnList)
        except KeyError:
            print('Add ISBN13 Column in the first sheet of excel, Skipping The File')

    def ReadCurrentISBN(self):
        for self.currentISBN in self.isbnList:
            self.bookCode = self.currentISBN
            self.currISBNRecord += 1
            self.currentISBN = str(self.currentISBN)
            self.resetISBNValues()
            self.CheckISBN()
            self.SetAmazonSearchLink()
            self.OpenAmazonPage()
            self.CheckPrimeStatus()
            self.GetAllValidLinks()
            self.GetCurrentLink()

    def ReadFileNames(self):
        self.all_excel_filenames = []
        for name in os.listdir(self.folderDirectory):
            if name[-5:] == '.xlsx' or name[-4:] == '.xls':
                self.all_excel_filenames.append(os.path.join(self.folderDirectory, name))
        for current_filename in self.all_excel_filenames:
            if current_filename:
                self.SetFileValues()
                self.CreateOutputFile(current_filename)
                self.ReadCurrentFile(current_filename)
                self.ReadISBN(current_filename)
                self.ReadCurrentISBN()

    def SetAmazonSearchLink(self):
        if self.isISBN:
            self.amazonSearchLink = 'https://www.amazon.in/s?i=stripbooks&rh=p_66%3A' + str(self.currentISBN)
        else:
            self.amazonSearchLink = 'https://www.amazon.in/s?k=' + str(self.currentISBN) + "&i=stripbooks"

    def SetupChromeDriver(self):
        self.options = Options()
        self.options.add_argument('--no-sandbox')
        self.options.add_experimental_option('excludeSwitches', ['enable-logging'])
        self.chrome_driver_path = r"chromedriver.exe"
        self.driver = webdriver.Chrome(executable_path=self.chrome_driver_path, options=self.options)

    def SetupFirefoxDriver(self):
        self.options = FirefoxOptions()
        self.options.add_argument('--no-sandbox')
        self.firefox_driver_path = r"geckodriver.exe"
        self.driver = webdriver.Firefox(executable_path=self.firefox_driver_path, options=self.options)

    def WriteRecord(self):
        if os.path.exists(self.outputFile):
            headers_row = [self.date, self.bookCode, self.currentISBN, self.coverType, self.prime, self.ranking,
                           self.buybox,
                           self.buyboxSeller,
                           self.buyboxPrice, self.reproListing,self.pages]
            output_workbook = load_workbook(self.outputFile)
            output_worksheet = output_workbook.worksheets[0]
            output_worksheet.append(headers_row)
            output_workbook.save(self.outputFile)


amazonScraper = AmazonScraper()
